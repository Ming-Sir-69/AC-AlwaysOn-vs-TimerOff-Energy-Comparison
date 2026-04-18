"""
Excel 数据记录模块

将仿真结果、主效应分析、方差分析导出为包含 7 个 Sheet 的 Excel 文件。
使用 openpyxl 实现，公式和条件格式均为原生 Excel 功能。
"""

import os
import sys
import pandas as pd
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from config import FACTORS_L25, EXCEL_STYLE


# ── 样式常量 ──
HEADER_FILL = PatternFill(start_color=EXCEL_STYLE['header_fill'],
                          end_color=EXCEL_STYLE['header_fill'], fill_type='solid')
HEADER_FONT = Font(name='Microsoft YaHei', bold=True,
                   color=EXCEL_STYLE['header_font_color'], size=11)
ZEBRA_FILL = PatternFill(start_color=EXCEL_STYLE['zebra_fill'],
                         end_color=EXCEL_STYLE['zebra_fill'], fill_type='solid')
WINNER_B_FILL = PatternFill(start_color=EXCEL_STYLE['winner_b_fill'],
                            end_color=EXCEL_STYLE['winner_b_fill'], fill_type='solid')
WINNER_A_FILL = PatternFill(start_color=EXCEL_STYLE['winner_a_fill'],
                            end_color=EXCEL_STYLE['winner_a_fill'], fill_type='solid')
CRITICAL_FILL = PatternFill(start_color=EXCEL_STYLE['critical_fill'],
                            end_color=EXCEL_STYLE['critical_fill'], fill_type='solid')
KPI_GREEN_FILL = PatternFill(start_color=EXCEL_STYLE['kpi_fill'],
                             end_color=EXCEL_STYLE['kpi_fill'], fill_type='solid')
KPI_RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='D9DEE7'),
    right=Side(style='thin', color='D9DEE7'),
    top=Side(style='thin', color='D9DEE7'),
    bottom=Side(style='thin', color='D9DEE7'),
)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


def _apply_header_style(ws, row, max_col):
    """为表头行应用深蓝底白字粗体样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _apply_border(ws, min_row, max_row, min_col, max_col):
    """为指定区域添加边框"""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def _auto_width(ws, min_col=1, max_col=None, min_width=8, max_width=30):
    """列宽自适应"""
    if max_col is None:
        max_col = ws.max_column
    for col in range(min_col, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    val_str = str(cell.value)
                    # 中文字符按 2 个宽度计算
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    max_len = max(max_len, char_len)
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _write_experiment_sheet(ws, df, sheet_name, highlight_critical=False):
    """
    写入实验结果 Sheet（L25 / L50 / 临界区加密共用）

    列映射：
        Run, t_off(h), T_out(C), T_in_init(C), T_set(C), tau(h), COP,
        E_A(kWh), E_B(kWh), delta_E(kWh), savings(%), winner
    """
    # 表头
    headers = [
        'Run', 't_off(h)', 'T_out(\u00b0C)', 'T_in_init(\u00b0C)',
        'T_set(\u00b0C)', 'tau(h)', 'COP',
        'E_A(kWh)', 'E_B(kWh)', 'delta_E(kWh)', 'savings(%)', 'winner'
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _apply_header_style(ws, 1, len(headers))

    # 数据行
    for i, (_, row_data) in enumerate(df.iterrows()):
        row_num = i + 2  # Excel 行号（从第 2 行开始）

        # 硬编码列：Run, t_off, T_out, T_in_init, T_set, tau, COP, E_A, E_B
        ws.cell(row=row_num, column=1, value=int(row_data['Run']))
        ws.cell(row=row_num, column=2, value=row_data['t_off'])
        ws.cell(row=row_num, column=3, value=row_data['T_out'])
        ws.cell(row=row_num, column=4, value=row_data['T_in_init'])
        ws.cell(row=row_num, column=5, value=row_data['T_set'])
        ws.cell(row=row_num, column=6, value=row_data['tau'])
        ws.cell(row=row_num, column=7, value=row_data['COP'])

        # E_A (H列), E_B (I列) — 仿真值硬编码
        ws.cell(row=row_num, column=8, value=row_data['E_A'])
        ws.cell(row=row_num, column=8).number_format = '0.0000'
        ws.cell(row=row_num, column=9, value=row_data['E_B'])
        ws.cell(row=row_num, column=9).number_format = '0.0000'

        # delta_E (J列) — Excel 公式: =H{row}-I{row}
        ws.cell(row=row_num, column=10).value = f'=H{row_num}-I{row_num}'
        ws.cell(row=row_num, column=10).number_format = '0.0000'

        # savings(%) (K列) — Excel 公式: =J{row}/H{row}*100
        ws.cell(row=row_num, column=11).value = f'=J{row_num}/H{row_num}*100'
        ws.cell(row=row_num, column=11).number_format = '0.00'

        # winner (L列) — Excel 公式: =IF(J{row}>0,"B关机更省","A常开更省")
        ws.cell(row=row_num, column=12).value = f'=IF(J{row_num}>0,"B关机更省","A常开更省")'

        # 居中对齐
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = CENTER_ALIGN

    # ── 条件格式：winner 列 ──
    last_row = len(df) + 1
    range_str = f'A2:L{last_row}'

    # "B关机更省" -> 整行浅红
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='equal',
                   formula=['"B关机更省"'],
                   fill=WINNER_B_FILL)
    )
    # "A常开更省" -> 整行浅蓝
    ws.conditional_formatting.add(
        range_str,
        CellIsRule(operator='equal',
                   formula=['"A常开更省"'],
                   fill=WINNER_A_FILL)
    )

    # ── 斑马纹（偶数行浅灰） ──
    for i in range(len(df)):
        row_num = i + 2
        if i % 2 == 1:  # 偶数行（0-indexed 的奇数 = Excel 的偶数行）
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = ZEBRA_FILL

    # ── 临界区加密额外高亮：|delta_E| < 0.5 标黄色 ──
    if highlight_critical:
        for i in range(len(df)):
            row_num = i + 2
            abs_delta = abs(df.iloc[i]['delta_E'])
            if abs_delta < 0.5:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = CRITICAL_FILL

    # ── savings(%) 列绿色数据条 ──
    ws.conditional_formatting.add(
        f'K2:K{last_row}',
        DataBarRule(
            start_type='min', end_type='max',
            color='63BE7B'  # 绿色
        )
    )

    # 边框
    _apply_border(ws, 1, last_row, 1, len(headers))

    # 列宽自适应
    _auto_width(ws, 1, len(headers))


def _write_config_sheet(ws):
    """Sheet 1: 实验参数配置"""
    headers = ['因子编号', '因子名称', '单位', '水平1', '水平2', '水平3', '水平4', '水平5']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _apply_header_style(ws, 1, len(headers))

    for i, (key, info) in enumerate(FACTORS_L25.items()):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=key)
        ws.cell(row=row_num, column=2, value=info['name'])
        ws.cell(row=row_num, column=3, value=info['unit'])
        for j, lv in enumerate(info['levels']):
            ws.cell(row=row_num, column=4 + j, value=lv)
        # 居中对齐
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = CENTER_ALIGN

    _apply_border(ws, 1, 1 + len(FACTORS_L25), 1, len(headers))
    _auto_width(ws, 1, len(headers))


def _write_main_effects_sheet(ws, main_effects):
    """Sheet 5: 主效应分析"""
    factor_names = {
        't_off': '出门时间(h)',
        'T_out': '室外温度(\u00b0C)',
        'T_in_init': '室内初始温(\u00b0C)',
        'T_set': '目标温度(\u00b0C)',
        'tau': '热惯性\u03c4(h)',
        'COP': 'COP',
    }

    current_row = 1
    for factor_col in ['t_off', 'T_out', 'T_in_init', 'T_set', 'tau', 'COP']:
        effects = main_effects.get(factor_col)
        if effects is None or len(effects) == 0:
            continue

        # 因子名称标题
        display_name = factor_names.get(factor_col, factor_col)
        ws.cell(row=current_row, column=1, value=display_name)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        # 表头
        ws.cell(row=current_row, column=1, value='水平值')
        ws.cell(row=current_row, column=2, value='\u0394E均值(kWh)')
        _apply_header_style(ws, current_row, 2)
        current_row += 1

        # 找到最大和最小值用于 KPI 高亮
        vals = effects.values
        max_val = max(vals)
        min_val = min(vals)

        for lv, mean_val in effects.items():
            ws.cell(row=current_row, column=1, value=lv)
            ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN
            ws.cell(row=current_row, column=2, value=round(mean_val, 4))
            ws.cell(row=current_row, column=2).number_format = '0.0000'
            ws.cell(row=current_row, column=2).alignment = CENTER_ALIGN

            # KPI 高亮
            if mean_val == max_val:
                ws.cell(row=current_row, column=2).fill = KPI_GREEN_FILL
            elif mean_val == min_val and min_val < 0:
                ws.cell(row=current_row, column=2).fill = KPI_RED_FILL

            ws.cell(row=current_row, column=1).border = THIN_BORDER
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            current_row += 1

        # 空行分隔
        current_row += 1

    _auto_width(ws, 1, 2)


def _write_anova_sheet(ws, anova_table):
    """Sheet 6: 方差分析"""
    headers = ['Factor', 'SS', 'df', 'MS', 'Contribution(%)']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _apply_header_style(ws, 1, len(headers))

    for i, (_, row_data) in enumerate(anova_table.iterrows()):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=row_data['Factor'])
        ws.cell(row=row_num, column=2, value=row_data['SS'])
        ws.cell(row=row_num, column=2).number_format = '0.0000'
        ws.cell(row=row_num, column=3, value=int(row_data['df']))
        ws.cell(row=row_num, column=4, value=row_data['MS'])
        ws.cell(row=row_num, column=4).number_format = '0.0000'
        ws.cell(row=row_num, column=5, value=row_data['Contribution(%)'])
        ws.cell(row=row_num, column=5).number_format = '0.00'

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = CENTER_ALIGN

        # Contribution(%) > 10% 标绿色加粗
        if row_data['Factor'] != 'Error' and row_data['Contribution(%)'] > 10:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = KPI_GREEN_FILL
                ws.cell(row=row_num, column=col).font = Font(bold=True)

    _apply_border(ws, 1, 1 + len(anova_table), 1, len(headers))
    _auto_width(ws, 1, len(headers))


def _write_critical_summary_sheet(ws):
    """Sheet 7: 临界点汇总"""
    headers = ['工况描述', '临界出门时长(h)', '说明']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    _apply_header_style(ws, 1, len(headers))

    # 预填典型工况的临界点数据
    typical_cases = [
        ('普通住宅, T_out=32\u00b0C, \u03c4=6h, COP=3.3', '~8.5',
         '大多数家庭典型工况，出门超过此时间关机更省'),
        ('保温良好, T_out=30\u00b0C, \u03c4=9h, COP=3.6', '~6.0',
         '保温好的房子临界时间更短'),
        ('保温差, T_out=36\u00b0C, \u03c4=2h, COP=2.6', '~12.0',
         '保温差+高温天气，需要更长时间关机才划算'),
        ('极端高温, T_out=40\u00b0C, \u03c4=4h, COP=3.0', '~10.0',
         '极端高温下回冷成本高，临界时间较长'),
        ('高效空调, T_out=34\u00b0C, \u03c4=6h, COP=4.0', '~5.5',
         '高COP空调回冷效率高，关机更划算'),
    ]

    for i, (desc, t_crit, note) in enumerate(typical_cases):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=desc)
        ws.cell(row=row_num, column=2, value=t_crit)
        ws.cell(row=row_num, column=3, value=note)
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).alignment = LEFT_ALIGN if col in (1, 3) else CENTER_ALIGN

    _apply_border(ws, 1, 1 + len(typical_cases), 1, len(headers))
    _auto_width(ws, 1, len(headers), min_width=15, max_width=55)


def export_all_results(df_l25, df_l50, df_cz, main_effects, anova_table, output_path):
    """
    导出包含 7 个 Sheet 的 Excel 文件

    参数:
        df_l25 (pd.DataFrame): L25 基础实验结果
        df_l50 (pd.DataFrame): L50 升级实验结果
        df_cz (pd.DataFrame): 临界区加密实验结果
        main_effects (dict): 主效应分析结果
        anova_table (pd.DataFrame): 方差分析表
        output_path (str): Excel 输出路径
    """
    wb = Workbook()

    # Sheet 1: 实验参数配置
    ws1 = wb.active
    ws1.title = '实验参数配置'
    _write_config_sheet(ws1)

    # Sheet 2: L25基础结果
    ws2 = wb.create_sheet('L25基础结果')
    _write_experiment_sheet(ws2, df_l25, 'L25基础结果', highlight_critical=False)

    # Sheet 3: L50升级结果
    ws3 = wb.create_sheet('L50升级结果')
    _write_experiment_sheet(ws3, df_l50, 'L50升级结果', highlight_critical=False)

    # Sheet 4: 临界区加密
    ws4 = wb.create_sheet('临界区加密')
    _write_experiment_sheet(ws4, df_cz, '临界区加密', highlight_critical=True)

    # Sheet 5: 主效应分析
    ws5 = wb.create_sheet('主效应分析')
    _write_main_effects_sheet(ws5, main_effects)

    # Sheet 6: 方差分析
    ws6 = wb.create_sheet('方差分析')
    _write_anova_sheet(ws6, anova_table)

    # Sheet 7: 临界点汇总
    ws7 = wb.create_sheet('临界点汇总')
    _write_critical_summary_sheet(ws7)

    # 确保输出目录存在
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"[Excel] 已保存: {output_path}")
